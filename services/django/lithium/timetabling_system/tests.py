from io import BytesIO
from unittest import mock

import pandas as pd
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import DatabaseError
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font

from timetabling_system.models import Venue


class HomePageTests(TestCase):
    def test_home_page_renders(self):
        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "timetabling_system/home.html")


class AboutPageTests(TestCase):
    def test_about_page_renders(self):
        response = self.client.get(reverse("about"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "timetabling_system/about.html")


class HealthTests(TestCase):
    def test_healthz_ok(self):
        response = self.client.get(reverse("healthz"))

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content,
            {"status": "ok", "services": {"database": {"status": "ok"}}},
        )

    def test_healthz_db_error_returns_503(self):
        cursor_mock = mock.MagicMock()
        cursor_mock.__enter__.side_effect = DatabaseError("boom")

        with mock.patch("timetabling_system.views.connection") as mocked_connection:
            mocked_connection.cursor.return_value = cursor_mock

            response = self.client.get(reverse("healthz"))

        self.assertEqual(response.status_code, 503)
        self.assertJSONEqual(
            response.content,
            {
                "status": "error",
                "services": {
                    "database": {"status": "error", "error": "boom"},
                },
            },
        )


class UploadTimetableFileTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="admin",
            email="admin@example.com",
            password="StrongPass123!",
            is_staff=True,
            is_superuser=True,
        )
        self.client.login(username="admin", password="StrongPass123!")

    def _build_excel_upload(self, rows):
        buffer = BytesIO()
        pd.DataFrame(rows).to_excel(buffer, index=False)
        buffer.seek(0)

        return SimpleUploadedFile(
            "test.xlsx",
            buffer.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    def _build_venue_upload(self):
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Monday"
        ws["B1"] = "Tuesday"
        ws["A2"] = "01/01/2025"
        ws["B2"] = "02/01/2025"
        ws["A3"] = "Hall A"
        ws["B3"] = "Lab B"
        ws["B3"].font = Font(color="FF0000")  # mark as not accessible

        wb.save(buffer)
        buffer.seek(0)

        return SimpleUploadedFile(
            "venues.xlsx",
            buffer.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    def test_upload_missing_file_returns_400(self):
        response = self.client.post(reverse("upload-exams"))
        self.assertEqual(response.status_code, 400)
        self.assertJSONEqual(
            response.content,
            {"status": "error", "message": "No file uploaded."},
        )

    def test_upload_exam_file_parses_successfully(self):
        upload = self._build_excel_upload(
            [
                {
                    "exam_code": "ABC123",
                    "exam_name": "Sample Exam",
                    "exam_date": "2024-10-01",
                    "exam_start": "09:00",
                    "exam_end": "12:00",
                    "exam_length": 180,
                    "exam_type": "Written",
                    "main_venue": "Main Hall",
                    "school": "Engineering",
                }
            ]
        )

        response = self.client.post(reverse("upload-exams"), {"file": upload})

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["type"], "Exam")
        self.assertEqual(payload["rows"][0]["exam_code"], "ABC123")

    def test_upload_venue_file_populates_database(self):
        upload = self._build_venue_upload()

        response = self.client.post(reverse("upload-exams"), {"file": upload})

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["type"], "Venue")
        self.assertEqual(payload["records_created"], 2)
        self.assertEqual(payload["records_updated"], 0)

        venues = {v.venue_name: v for v in Venue.objects.all()}
        self.assertEqual(len(venues), 2)
        self.assertTrue(venues["Hall A"].is_accessible)
        self.assertFalse(venues["Lab B"].is_accessible)
        self.assertEqual(venues["Hall A"].capacity, 0)
        self.assertEqual(venues["Hall A"].venuetype, "school_to_sort")

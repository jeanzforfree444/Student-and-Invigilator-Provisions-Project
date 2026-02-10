from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def _cell_to_date_text(cell):
    val = cell.value
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)):
        try:
            return from_excel(val).date().isoformat()
        except Exception:
            return str(val)
    if val:
        try:
            parsed = datetime.fromisoformat(str(val))
            return parsed.date().isoformat()
        except Exception:
            return str(val).strip()
    return None

def parse_venue_file(file):
    wb = load_workbook(file)
    ws = wb.active
    results = []
    venue_index = {}  # venue_name -> accessibility flag (False if any instance is inaccessible)
    # Find the first non-empty row; some templates start with a blank row.
    header_row = None
    for row in range(1, ws.max_row + 1):
        if any(ws.cell(row, col).value for col in range(1, ws.max_column + 1)):
            header_row = row
            break
    if header_row is None or header_row + 1 > ws.max_row:
        return {
            "status": "error",
            "type": "Venue",
            "message": "Could not locate header rows in venue file."
        }

    # Read column pairs: (header row, date row, data rows)
    date_row = header_row + 1
    first_data_row = date_row + 1

    for col in range(1, ws.max_column + 1):
        day_cell = ws.cell(header_row, col)
        date_cell = ws.cell(date_row, col)

        day_text = str(day_cell.value).strip() if day_cell.value else None
        date_text = _cell_to_date_text(date_cell)

        # Skip empty columns
        if not day_text:
            continue

        rooms = []

        for row in range(first_data_row, ws.max_row + 1):
            cell = ws.cell(row, col)
            value = cell.value

            if not value: continue

            # Detect red font (non-accessible)
            font_color = cell.font.color
            rgb = str(font_color.rgb).upper() if font_color and font_color.rgb else ""
            is_red = (
                font_color
                and font_color.type == "rgb"
                and "FF0000" in rgb
            )

            room_name = str(value).strip()
            accessible = not is_red

            rooms.append({
                "name": room_name,
                "accessible": accessible
            })

            # Track venue-level accessibility; once false, remain false.
            venue_index[room_name] = venue_index.get(room_name, True) and accessible
        results.append({
            "day": day_text,
            "date": date_text,
            "rooms": rooms
        })

    venues = [
        {"name": name, "is_accessible": accessible}
        for name, accessible in venue_index.items()
    ]

    return {
        "status": "ok",
        "type": "Venue",
        "days": results,
        "venues": venues,
    }

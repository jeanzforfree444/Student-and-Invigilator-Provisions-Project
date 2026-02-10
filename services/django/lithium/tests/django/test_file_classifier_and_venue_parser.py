import pandas as pd
from datetime import datetime, date
from django.test import TestCase

from timetabling_system.utils import file_classifier
from timetabling_system.utils.venue_parser import parse_venue_file, _cell_to_date_text


class FileClassifierTests(TestCase):
    def test_detect_exam_vs_provision(self):
        exam_df = pd.DataFrame(columns=["Exam Code", "Exam Name", "Exam date", "Main Venue"])
        self.assertTrue(file_classifier.detect_exam_file(exam_df))
        self.assertFalse(file_classifier.detect_provision_file(exam_df))

        prov_df = pd.DataFrame(columns=["Student ID", "Student Name", "Exam provision data as presented to registry"])
        self.assertTrue(file_classifier.detect_provision_file(prov_df))
        self.assertFalse(file_classifier.detect_exam_file(prov_df))

    def test_detect_venue_file_prefers_exam_and_provision_checks(self):
        df = pd.DataFrame(columns=["Exam Code", "Exam Name", "Exam date", "Main Venue"])
        self.assertFalse(file_classifier.detect_venue_file(df))

        # Column-based weekday/date layout should be detected as venue
        venue_df = pd.DataFrame(
            [
                ["Monday", "Tuesday"],
                [45866, 45867],
                ["Room A", "Room B"],
            ]
        )
        self.assertTrue(file_classifier.detect_venue_file(venue_df))

    def test_canonical_columns_uses_fallback_header_row(self):
        # Leading unnamed columns; real header is in the first data row
        df = pd.DataFrame(
            [["Exam Code", "Exam Name", "Exam date"], ["CHEM101", "Chemistry", "2025-06-01"]],
            columns=["Unnamed: 0", "Unnamed: 1", "Unnamed: 2"],
        )
        canonical = file_classifier._canonical_columns(df)
        self.assertIn("exam_code", canonical)
        self.assertIn("exam_name", canonical)

    def test_normalized_columns_and_header_error_fallback(self):
        df = pd.DataFrame(columns=["Exam Code", "Exam Name"])
        self.assertEqual(
            file_classifier._normalized_columns(df),
            {"exam_code", "exam_name"},
        )

        class BoomDF:
            columns = ["Exam Code"]
            index = [0]

            @property
            def iloc(self):
                raise RuntimeError("boom")

        canonical = file_classifier._canonical_columns(BoomDF())
        self.assertIn("exam_code", canonical)

    def test_looks_like_date_cell_various_inputs(self):
        self.assertTrue(file_classifier._looks_like_date_cell(datetime.now()))
        self.assertTrue(file_classifier._looks_like_date_cell("2024-12-01"))
        self.assertFalse(file_classifier._looks_like_date_cell(None))
        self.assertFalse(file_classifier._looks_like_date_cell(""))
        self.assertFalse(file_classifier._looks_like_date_cell(float("nan")))
        self.assertTrue(file_classifier._looks_like_date_cell("12345"))
        self.assertFalse(file_classifier._looks_like_date_cell("123"))
        # Value that cannot be coerced to float hits the exception path
        class NoFloat:
            def __str__(self):
                return "not_a_number"
        self.assertFalse(file_classifier._looks_like_date_cell(NoFloat()))

    def test_detect_venue_file_handles_invalid_input(self):
        self.assertFalse(file_classifier.detect_venue_file(None))


class VenueParserTests(TestCase):
    def test_cell_to_date_text_handles_excel_serial_and_strings(self):
        class DummyCell:
            def __init__(self, value):
                self.value = value
        self.assertEqual(_cell_to_date_text(DummyCell(45866)), "2025-07-28")
        self.assertEqual(_cell_to_date_text(DummyCell("2025-07-29")), "2025-07-29")
        self.assertIsNone(_cell_to_date_text(DummyCell(None)))
        # Non-date string falls back to stripped text
        self.assertEqual(_cell_to_date_text(DummyCell("notadate")), "notadate")

    def test_cell_to_date_text_handles_datetime_and_invalid_serial(self):
        class DummyCell:
            def __init__(self, value):
                self.value = value
        now = datetime(2025, 7, 30, 10, 0)
        self.assertEqual(_cell_to_date_text(DummyCell(now)), now.date().isoformat())
        self.assertEqual(_cell_to_date_text(DummyCell(date(2025, 7, 31))), "2025-07-31")
        # Invalid excel serial falls back to string
        self.assertEqual(_cell_to_date_text(DummyCell(float("nan"))), "nan")

    def test_parse_venue_file_errors_on_empty_sheet(self):
        from openpyxl import Workbook
        from tempfile import NamedTemporaryFile

        wb = Workbook()
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            result = parse_venue_file(tmp.name)
        self.assertEqual(result["status"], "error")

    def test_parse_venue_file_collects_accessibility(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from tempfile import NamedTemporaryFile

        wb = Workbook()
        ws = wb.active
        ws.append(["Monday"])
        ws.append(["2025-07-28"])
        ws.append(["Room A"])
        ws.append(["Room B"])
        # Mark Room B as red (inaccessible)
        ws["A4"].font = Font(color="FF0000")

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            result = parse_venue_file(tmp.name)

        self.assertEqual(result["status"], "ok")
        venues = {v["name"]: v for v in result["venues"]}
        self.assertTrue(venues["Room A"]["is_accessible"])
        self.assertFalse(venues["Room B"]["is_accessible"])

    def test_parse_venue_file_skips_empty_header_columns(self):
        from openpyxl import Workbook
        from tempfile import NamedTemporaryFile

        wb = Workbook()
        ws = wb.active
        ws.append([None, "Tuesday"])
        ws.append([None, "2025-07-29"])
        ws.append([None, "Room B"])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            result = parse_venue_file(tmp.name)

        self.assertEqual(result["status"], "ok")
        self.assertEqual(len(result["days"]), 1)
